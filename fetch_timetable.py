"""
Timetable Fetcher and Parser

This script fetches, parses, and normalizes university timetable schedules 
from three public Google Sheets into a unified `timetable.json` file.
"""
import json
import logging
import requests
import io
import re
import time
import openpyxl
from typing import List, Dict, Any
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

URLS = {
    "FSC": "https://docs.google.com/spreadsheets/d/1vlTuotLw34fedME3gNQj09cZw-todVomxAiu5P1wZ6Q/htmlview?gid=1174567785",
    "FSM": "https://docs.google.com/spreadsheets/d/1AnFQQhv9lu4grESE2ypbDG7E1QOPGgGCRiejem5ocPw/export?format=xlsx&gid=0",
    "FSE": "https://docs.google.com/spreadsheets/d/1fL2TWhPgbPc2d66vm_KywTpdsGBIaBLqlmz4JLPudCw/export?format=xlsx&gid=115356958"
}

# Example Color mappings for FSC - Configure according to the actual sheets
# Note: Color hex codes need to be updated from openpyxl's ARGB format to standard CSS hex/RGB formats to match the HTML view.
COLOR_LEGEND = {
    "FFFF0000": {"department": "CS", "degree": "BS", "batch": "2026", "semester": "1"}, 
    "FF00FF00": {"department": "DS", "degree": "BS", "batch": "2025", "semester": "3"},
    "FF0000FF": {"department": "AI", "degree": "BS", "batch": "2024", "semester": "5"},
    "FFFFFF00": {"department": "SE", "degree": "BS", "batch": "2023", "semester": "7"}
}

FSM_COLOR_LEGEND = {
    "FFCC99FF": {"department": "AF", "degree": "BS", "batch": "2026"},
    "FFD9EAD3": {"department": "AF", "degree": "BS", "batch": "2025"},
    "FFFF89D8": {"department": "BBA", "degree": "BBA", "batch": "2026"},
    "FFFFB265": {"department": "BBA", "degree": "BBA", "batch": "2025"},
    "FFF28E86": {"department": "BA", "degree": "BS", "batch": "2026"},
    "FF00FFFF": {"department": "BA", "degree": "BS", "batch": "2025"},
    "FFFCD6EC": {"department": "FT", "degree": "BS", "batch": "2026"},
    "FFBC8E03": {"department": "FT", "degree": "BS", "batch": "2025"},
    "FFB5E3E8": {"department": "AF", "degree": "BS", "batch": "2024"},
    "FFFEF2CD": {"department": "FT", "degree": "BS", "batch": "2023"},
    "FFEA4335": {"department": "BBA", "degree": "BBA", "batch": "2024"},
    "FF79BCFF": {"department": "BBA", "degree": "BBA", "batch": "2023"},
    "FFA6E3B7": {"department": "BA", "degree": "BS", "batch": "2024"},
    "FF2F9299": {"department": "BA", "degree": "BS", "batch": "2023"},
    "FFFFE1CC": {"department": "FT", "degree": "BS", "batch": "2024"},
    "FF993366": {"department": "AF", "degree": "BS", "batch": "2023"}
}

FSE_COLOR_LEGEND = {
    "FFFFC000": {"department": "EE", "degree": "BS", "batch": "2026"},
    "FFFBBC04": {"department": "EE", "degree": "BS", "batch": "2026"},
    "FF00B0F0": {"department": "EE", "degree": "BS", "batch": "2025"},
    "FF00B050": {"department": "EE", "degree": "BS", "batch": "2024"},
    "FFEA4335": {"department": "EE", "degree": "BS", "batch": "2023"},
    "FFEF91F1": {"department": "CE", "degree": "BS", "batch": "2026"},
    "FFC0D91E": {"department": "CE", "degree": "BS", "batch": "2025"},
    "FFF4B084": {"department": "CE", "degree": "BS", "batch": "2024"}
}

# School of Computing Color Legend (Batch mapping only)
FSC_COLOR_LEGEND = {
    # BS CS
    "FFB740": "2026",
    "6D5200": "2025",
    "C39401": "2024",
    "FFE599": "2023",
    # BS DS
    "7F4CFF": "2026",
    "351C75": "2025",
    "B17FD7": "2024",
    "B4A7D6": "2023",
    # BS AI
    "00F600": "2026",
    "274E13": "2025",
    "6AA84F": "2024",
    "B6D7A8": "2023",
    # BS CY
    "0000FF": "2026",
    "073763": "2025",
    "599DDA": "2024",
    "ABCCEB": "2023",
    # BS SE
    "E62C06": "2026",
    "85200C": "2025",
    "DD7E6B": "2024",
    "F4CCCC": "2023"
}

def clean_text(text: Any) -> str:
    if not text:
        return ""
    return str(text).strip()

def normalize_time(t: str) -> str:
    """Converts university times (1-7 are PM) to 24-hour format for correct sorting."""
    if not t: return ""
    t_clean = t.replace(' ', '').replace('AM', '').replace('PM', '').strip()
    parts = t_clean.split(':')
    if len(parts) != 2: return t
    try:
        h = int(parts[0])
        m = int(parts[1][:2]) # in case there is trailing text
        # If hour is between 1 and 7 (inclusive), it is PM in university context.
        if 1 <= h <= 7:
            h += 12
        return f"{h:02d}:{m:02d}"
    except ValueError:
        return t

def download_workbook(url: str) -> openpyxl.Workbook:
    """Downloads Google Sheet as an Excel file and loads into openpyxl"""
    # Cache-busting parameter
    cb_url = f"{url}&_cb={int(time.time())}" if "?" in url else f"{url}?_cb={int(time.time())}"
    logging.info(f"Downloading data from {cb_url}")
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Cache-Control': 'no-cache, no-store, must-revalidate',
        'Pragma': 'no-cache',
        'Expires': '0'
    }
    response = requests.get(cb_url, headers=headers, timeout=30)
    response.raise_for_status()
    return openpyxl.load_workbook(filename=io.BytesIO(response.content), data_only=True)

def get_timetable_sheet(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    """Finds the correct sheet containing the timetable."""
    for name in wb.sheetnames:
        lower_name = name.lower()
        if 'timetable' in lower_name or 'schedule' in lower_name:
            return wb[name]
    return wb.active

def generate_rag_summary(school: str, dept: str, degree: str, batch: str, section: str, 
                         course: str, room: str, day: str, t_start: str, t_end: str, is_lab: bool, is_rescheduled: bool = False, is_repeat: bool = False, is_cancelled: bool = False) -> str:
    """Generates a text summary string suitable for RAG ingestion."""
    lab_text = " (Lab)" if is_lab else ""
    status_prefix = ""
    if is_cancelled:
        status_prefix = "[CANCELLED] "
    elif is_rescheduled:
        status_prefix = "[RESCHEDULED] "
        
    repeat_tag = " [REPEAT COURSE]" if is_repeat else ""
    return f"{status_prefix}{degree} {dept} (Batch {batch}, Section {section}) has {course}{lab_text}{repeat_tag} in Room {room} on {day} from {t_start} to {t_end}."

def extract_time_slots(sheet: openpyxl.worksheet.worksheet.Worksheet, start_col: int = 2) -> List[tuple]:
    """Extracts column mappings for time slots by examining headers in the first 10 rows."""
    time_slots = []
    for r in range(1, 11):
        for col_idx in range(start_col, sheet.max_column + 1):
            val = clean_text(sheet.cell(row=r, column=col_idx).value)
            # Identify typical time slot format like '08:30-09:50'
            if val and "-" in val and any(char.isdigit() for char in val):
                time_slots.append((col_idx, val))
        if time_slots:
            break # Found the header row
    return time_slots

# ── FSC (School of Computing) constants ────────────────────────────────────────
FSC_SPREADSHEET_ID = "1vlTuotLw34fedME3gNQj09cZw-todVomxAiu5P1wZ6Q"

FSC_DAY_GIDS = {
    "Monday":    "1882612924",
    "Tuesday":   "945396749",
    "Wednesday": "542677125",
    "Thursday":  "571927841",
    "Friday":    "1783333514",
    "Saturday":  "1949393871",
}

FSC_DEPT_MAP = {
    "CS": "CS", "DS": "DS", "AI": "AI",
    "CY": "CY", "SE": "SE",
}

def fetch_fsc_gids() -> Dict[str, str]:
    """Dynamically fetches GIDs for FSC timetable since they might change when updated."""
    url = f"https://docs.google.com/spreadsheets/d/{FSC_SPREADSHEET_ID}/htmlview?_cb={int(time.time())}"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Cache-Control': 'no-cache, no-store, must-revalidate'
    }
    try:
        html = requests.get(url, headers=headers, timeout=30).text
        matches = re.findall(r'items\.push\((.*?)\);', html)
        gids = {}
        for m in matches:
            name_match = re.search(r'name:\s*"([^"]+)"', m)
            gid_match = re.search(r'gid:\s*"(\d+)"', m)
            if name_match and gid_match:
                name = name_match.group(1).strip()
                gid = gid_match.group(1)
                name_lower = name.lower()
                if 'monday' in name_lower or 'tuesday' in name_lower or 'wednesday' in name_lower or \
                   'thursday' in name_lower or 'friday' in name_lower or 'saturday' in name_lower or 'sunday' in name_lower:
                    gids[name] = gid
        return gids
    except Exception as e:
        logging.error(f"Failed to fetch dynamic FSC GIDs: {e}")
        return {}

def parse_fsc() -> List[Dict[str, Any]]:
    """Parser for School of Computing — fetches each day's HTML frame by GID."""
    entries = []
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(
                extra_http_headers={
                    "Cache-Control": "no-cache, no-store, must-revalidate",
                    "Pragma": "no-cache"
                }
            )

            dynamic_gids = fetch_fsc_gids()
            day_gids_to_use = dynamic_gids if dynamic_gids else FSC_DAY_GIDS
            logging.info(f"Using FSC GIDs: {day_gids_to_use}")

            for day_name, gid in day_gids_to_use.items():
                frame_url = (
                    f"https://docs.google.com/spreadsheets/d/{FSC_SPREADSHEET_ID}"
                    f"/htmlview/sheet?headers=true&gid={gid}&_cb={int(time.time())}"
                )
                logging.info(f"Fetching FSC {day_name} (gid={gid})")

                try:
                    page = context.new_page()
                    page.goto(frame_url, wait_until="networkidle", timeout=60000)
                    page.wait_for_timeout(1500)
                    html_content = page.content()
                    page.close()
                except Exception as e:
                    logging.warning(f"Failed to load FSC {day_name}: {e}")
                    continue

                # Extract class to color mapping from style tags
                class_to_color = {}
                for match in re.finditer(r'\.(s\d+)\s*\{[^\}]*background-color:\s*(#[0-9a-fA-F]{6})', html_content):
                    class_to_color[match.group(1)] = match.group(2).upper().replace("#", "")

                soup = BeautifulSoup(html_content, "html.parser")
                table = soup.find("table")
                if not table:
                    logging.warning(f"No table for FSC {day_name}")
                    continue

                all_rows = table.find_all("tr")

                def extract_row_cells(r):
                    cells = r.find_all(["td", "th"])
                    grid = {}
                    vcol = 0
                    for c in cells:
                        colspan = int(c.get('colspan', 1))
                        grid[vcol] = (clean_text(c.get_text()), c, colspan)
                        vcol += colspan
                    return grid

                # Step 1: find time-slot header row
                time_col_map: Dict[int, str] = {}
                header_row_idx = None
                for ridx, row in enumerate(all_rows):
                    grid = extract_row_cells(row)
                    texts = [v[0] for v in grid.values()]
                    flat = " ".join(texts).lower()
                    if "room" in flat and "time" in flat:
                        for vcol, (txt, _, _) in grid.items():
                            if re.match(r"\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2}", txt):
                                time_col_map[vcol] = txt
                        header_row_idx = ridx
                        break

                if not time_col_map or header_row_idx is None:
                    logging.warning(f"Could not find time header for FSC {day_name}")
                    continue

                # Step 2: iterate data rows
                day_count = 0
                for row in all_rows[header_row_idx + 1:]:
                    grid = extract_row_cells(row)
                    room = ""
                    if 1 in grid and grid[1][0]:
                        room = grid[1][0]

                    if not room or len(room) > 25:
                        continue
                    
                    room_lower = room.lower()
                    if any(kw in room_lower for kw in ["bs ", "ms ", "phd"]):
                        room = "Unknown"
                    elif any(kw in room_lower for kw in [
                        "room", "time", "monday", "tuesday", "wednesday",
                        "thursday", "friday", "saturday", "sunday"
                    ]):
                        continue

                    # Step 3: each cell in the row
                    for vcol, (val, cell, colspan) in grid.items():
                        if not val or vcol == 0 or vcol == 1: # skip empty or row header cells
                            continue
                            
                        # Find closest time header
                        start_vcol = max((k for k in time_col_map.keys() if k <= vcol), default=None)
                        if start_vcol is None:
                            continue
                        time_val = time_col_map[start_vcol]

                        # Format: "Course Name (DEPT-Section)" e.g. "PF (CS-A)" or "OOP (CS-B, 25)" or "Seerah (C5-B)"
                        course_match = re.match(r"^(.+?)\s*\(([A-Z0-9]{2,3}-[A-Z0-9]+)(?:,\s*(\d+))?\)", val)
                        if not course_match:
                            continue

                        course_name = course_match.group(1).strip()
                        section_code = course_match.group(2).strip()
                        explicit_batch_code = course_match.group(3)
                        
                        is_rescheduled = "resch" in val.lower()
                        if is_rescheduled:
                            course_name = re.sub(r'(?i)\s*[-]*\s*resch', '', course_name).strip()

                        is_cancelled = "cancelled" in val.lower()
                        if is_cancelled:
                            course_name = re.sub(r'(?i)\s*[-]*\s*cancelled', '', course_name).strip()

                        # Skip postgraduate
                        if any(pg in val for pg in ["MS", "PhD", "PCS", "Repeat"]):
                            continue

                        dept_key = section_code.split("-")[0]
                        dept = FSC_DEPT_MAP.get(dept_key, dept_key)
                        
                        td_classes = cell.get("class", [])
                        cell_color = None
                        for c in td_classes:
                            if c in class_to_color:
                                cell_color = class_to_color[c]
                                break

                        batch = FSC_COLOR_LEGEND.get(cell_color, "Unknown") if cell_color else "Unknown"
                        is_repeat = (cell_color == "FFFF00")
                        
                        if explicit_batch_code:
                            explicit_b = explicit_batch_code.strip()
                            if len(explicit_b) == 2:
                                batch = "20" + explicit_b
                            else:
                                batch = explicit_b

                        # calculate exact time range based on colspan
                        t_parts = time_val.split("-")
                        t_start = normalize_time(t_parts[0].strip()) if t_parts else ""
                        
                        end_vcol = max((k for k in time_col_map.keys() if k < vcol + colspan), default=vcol)
                        end_time_val = time_col_map[end_vcol]
                        end_t_parts = end_time_val.split("-")
                        t_end = normalize_time(end_t_parts[1].strip()) if len(end_t_parts) > 1 else ""

                        # Override with explicit time if present in the raw cell value
                        explicit_time_match = re.search(r"(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})", val)
                        if explicit_time_match:
                            t_start = normalize_time(explicit_time_match.group(1).strip())
                            t_end = normalize_time(explicit_time_match.group(2).strip())
                            # Remove the explicit time from the course name
                            course_name = re.sub(r"\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2}", "", course_name).strip()
                        else:
                            # Fallback to predefined special course durations if no explicit time is found
                            special_durations = {
                                "seerah": 55,
                                "uhq-i&ii": 110,
                                "uhq-i & ii": 110,
                                "ideology": 105,
                                "islamic": 105,
                                "func eng": 105,
                                "uhq-ii": 55,
                                "arts & humanities": 105
                            }
                            c_lower = course_name.lower()
                            duration = next((v for k, v in special_durations.items() if k in c_lower), None)
                            if duration and t_start:
                                try:
                                    from datetime import datetime, timedelta
                                    dt = datetime.strptime(t_start, "%H:%M")
                                    dt += timedelta(minutes=duration)
                                    t_end = dt.strftime("%H:%M")
                                except:
                                    pass

                        is_lab = "lab" in course_name.lower() or "lab" in room.lower()
                        entry_id = f"FSC-{day_name[:3].upper()}-{room.replace('-','')}-{t_start.replace(':','')}-{section_code.replace('-','')}"

                        summary = generate_rag_summary(
                            "School of Computing", dept, "BS", batch,
                            section_code, course_name, room, day_name, t_start, t_end, is_lab, is_rescheduled, is_repeat, is_cancelled
                        )

                        entries.append({
                            "id": entry_id,
                            "school": "School of Computing",
                            "department": dept,
                            "degree": "BS",
                            "batch": batch,
                            "semester": "Unknown",
                            "course_name": course_name,
                            "section": section_code,
                            "instructor": None,
                            "room": room,
                            "day": day_name,
                            "time_start": t_start,
                            "time_end": t_end,
                            "is_lab": is_lab,
                            "is_rescheduled": is_rescheduled,
                            "is_repeat": is_repeat,
                            "is_cancelled": is_cancelled,
                            "rag_summary": summary,
                        })
                        day_count += 1

                logging.info(f"  FSC {day_name}: {day_count} entries")

            browser.close()

    except Exception as e:
        logging.error(f"Error parsing FSC: {e}", exc_info=True)

    logging.info(f"FSC total: {len(entries)} entries.")
    return entries

def parse_fsm() -> List[Dict[str, Any]]:
    """Parser for School of Management"""
    entries = []
    try:
        wb = download_workbook(URLS["FSM"])
        sheet = get_timetable_sheet(wb)
        time_slots = extract_time_slots(sheet, start_col=4)
        current_day = "Monday"
        
        for row_idx in range(4, sheet.max_row + 1):
            cell_A = clean_text(sheet.cell(row=row_idx, column=1).value)
            
            # Check for day block start (allow dynamic names like "Saturday (Sep 05)")
            cell_lower = cell_A.lower()
            if len(cell_A) < 40 and any(cell_lower.startswith(d) for d in ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]):
                current_day = cell_A.strip()
                
            room = clean_text(sheet.cell(row=row_idx, column=3).value)
            if not room:
                room = clean_text(sheet.cell(row=row_idx, column=2).value)
            if not room:
                continue
                
            for col_idx in range(4, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                val = clean_text(cell.value)
                if not val:
                    continue
                
                # Skip if this cell is purely a section tag
                if re.match(r'^([A-Z]{2,4})(\d{2})([A-Z/0-9]+)$', val.replace(' ', '')):
                    continue
                
                start_col = max((k for k, v in time_slots if k <= col_idx), default=None)
                if start_col is None:
                    continue
                time_val = next(v for k, v in time_slots if k == start_col)
                    
                color_hex = str(cell.fill.start_color.index).upper() if cell.fill else "Unknown"
                color_info = {"department": "Unknown", "degree": "BS", "batch": "Unknown"}
                
                # Check for color in legend, allow partial match
                for key in FSM_COLOR_LEGEND:
                    if key in color_hex or color_hex.replace("#", "") in key:
                        color_info = FSM_COLOR_LEGEND[key]
                        break

                dept_code = color_info["department"]
                batch = color_info["batch"]
                degree = color_info["degree"]
                semester = "Unknown"
                section = "Unknown"
                
                t_parts = time_val.split("-")
                t_start = normalize_time(t_parts[0].strip()) if len(t_parts) > 0 else ""
                t_end = normalize_time(t_parts[1].strip()) if len(t_parts) > 1 else ""
                
                explicit_time_match = re.search(r"\(?(\d{1,2}:\d{2})\s*(?:-|to)\s*(\d{1,2}:\d{2}(?:[ap]m)?)\)?", val, re.IGNORECASE)
                if explicit_time_match:
                    t_start = normalize_time(explicit_time_match.group(1).strip())
                    raw_t_end = explicit_time_match.group(2).strip().lower().replace("pm", "").replace("am", "")
                    t_end = normalize_time(raw_t_end)
                    course_name = re.sub(r"\(?\d{1,2}:\d{2}\s*(?:-|to)\s*\d{1,2}:\d{2}(?:[ap]m)?\)?", "", val, flags=re.IGNORECASE).strip()
                else:
                    course_name = val.strip()
                
                course_name = re.sub(r'[\n\r]+', ' ', course_name).strip()
                
                is_rescheduled = "resch" in val.lower()
                if is_rescheduled:
                    course_name = re.sub(r'(?i)\s*[-]*\s*resch', '', course_name).strip()
                
                is_cancelled = "cancelled" in val.lower()
                if is_cancelled:
                    course_name = re.sub(r'(?i)\s*[-]*\s*cancelled', '', course_name).strip()
                
                # Search forward for the next non-empty cell in the row
                for c in range(col_idx + 1, sheet.max_column + 1):
                    next_val = clean_text(sheet.cell(row=row_idx, column=c).value)
                    if next_val:
                        sec_match = re.search(r'^([A-Z]{2,4})(\d{2})([A-Z/0-9]+)$', next_val.replace(' ', ''))
                        if sec_match:
                            parsed_dept = sec_match.group(1)
                            sem_code = sec_match.group(2)
                            section = next_val.replace(' ', '')
                            
                            # Deduce department and degree
                            if parsed_dept == 'BSBA': dept_code = 'BA'
                            else: dept_code = parsed_dept
                            
                            if dept_code == 'BBA': degree = 'BBA'
                            else: degree = 'BS'
                            
                            # Deduce batch
                            if sem_code == '01': batch = '2026'
                            elif sem_code == '03': batch = '2025'
                            elif sem_code == '05': batch = '2024'
                            elif sem_code == '07': batch = '2023'
                        break
                
                school = "School of Management"
                is_lab = False
                
                summary = generate_rag_summary(school, dept_code, degree, batch, section, course_name, room, current_day, t_start, t_end, is_lab, is_rescheduled, False, is_cancelled)
                entry_id = f"FSM-{current_day[:3].upper()}-{room.replace('-', '')}-{t_start.replace(':', '')}"
                
                entries.append({
                    "id": entry_id,
                    "school": school,
                    "department": dept_code,
                    "degree": degree,
                    "batch": batch,
                    "semester": semester,
                    "course_name": course_name,
                    "section": section,
                    "instructor": None,
                    "room": room,
                    "day": current_day,
                    "time_start": t_start,
                    "time_end": t_end,
                    "is_lab": is_lab,
                    "is_rescheduled": is_rescheduled,
                    "is_repeat": False,
                    "is_cancelled": is_cancelled,
                    "rag_summary": summary
                })
    except Exception as e:
        logging.error(f"Error parsing FSM: {e}")
    return entries

def parse_fse() -> List[Dict[str, Any]]:
    """Parser for School of Engineering"""
    entries = []
    try:
        wb = download_workbook(URLS["FSE"])
        sheet = get_timetable_sheet(wb)
        
        # In FSE time slots might be offset if room numbers occupy 2 columns
        # extract_time_slots iterates from column 2 onwards, so it should catch them
        time_slots = extract_time_slots(sheet, start_col=2)
        current_day = "Monday"
        
        for row_idx in range(4, sheet.max_row + 1):
            cell_A = clean_text(sheet.cell(row=row_idx, column=1).value)
            cell_lower = cell_A.lower()
            if len(cell_A) < 40 and any(cell_lower.startswith(d) for d in ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]):
                current_day = cell_A.strip()
                
            # Room is often in col 3 or col 2
            room = clean_text(sheet.cell(row=row_idx, column=3).value)
            if not room:
                room = clean_text(sheet.cell(row=row_idx, column=2).value)
            if not room:
                continue
                
            for col_idx, time_val in time_slots:
                cell = sheet.cell(row=row_idx, column=col_idx)
                val = clean_text(cell.value)
                if not val:
                    continue
                    
                color_hex = str(cell.fill.start_color.index).upper() if cell.fill else "Unknown"
                color_info = {"department": "Unknown", "degree": "BS", "batch": "Unknown"}
                
                for key in FSE_COLOR_LEGEND:
                    if key in color_hex or color_hex.replace("#", "") in key:
                        color_info = FSE_COLOR_LEGEND[key]
                        break
                        
                batch = color_info["batch"]
                degree = color_info["degree"]
                    
                # Look ahead for instructor
                instructor = None
                next_val = clean_text(sheet.cell(row=row_idx + 1, column=col_idx).value)
                if next_val and "Dr." in next_val or "Ms." in next_val or "Mr." in next_val or "Engr." in next_val:
                    instructor = next_val
                
                # Newline separated for Course Title + Section, or split across rows
                lines = val.split('\n')
                if len(lines) >= 1:
                    first_line = lines[0].strip()
                    instructor = lines[1].strip() if len(lines) > 1 else None
                    
                    course_match = re.match(r"(.*?)\s+([A-Z]+-[A-Z]+|\b[A-Z]\b)$", first_line)
                    if course_match:
                        course_name = course_match.group(1).strip()
                        raw_section = course_match.group(2).strip()
                        if "-" in raw_section:
                            section = raw_section
                            dept = section.split('-')[0]
                        else:
                            dept = color_info["department"] if color_info["department"] != "Unknown" else "EE"
                            section = f"{dept}-{raw_section}"
                    else:
                        course_name = first_line
                        section = "Unknown"
                        dept = color_info["department"] if color_info["department"] != "Unknown" else "EE"
                        
                    is_rescheduled = "resch" in val.lower()
                    if is_rescheduled:
                        course_name = re.sub(r'(?i)\s*[-]*\s*resch', '', course_name).strip()
                        
                    is_cancelled = "cancelled" in val.lower()
                    if is_cancelled:
                        course_name = re.sub(r'(?i)\s*[-]*\s*cancelled', '', course_name).strip()
                        
                    t_parts = time_val.split("-")
                    t_start = normalize_time(t_parts[0].strip()) if len(t_parts) > 0 else ""
                    t_end = normalize_time(t_parts[1].strip()) if len(t_parts) > 1 else ""
                    
                    school = "School of Engineering"
                    semester = "Unknown"
                    is_lab = "lab" in course_name.lower() or "lab" in room.lower()
                    
                    summary = generate_rag_summary(school, dept, degree, batch, section, course_name, room, current_day, t_start, t_end, is_lab, is_rescheduled, False, is_cancelled)
                    entry_id = f"FSE-{current_day[:3].upper()}-{room.replace('-', '')}-{t_start.replace(':', '')}"
                    
                    entries.append({
                        "id": entry_id,
                        "school": school,
                        "department": dept,
                        "degree": degree,
                        "batch": batch,
                        "semester": semester,
                        "course_name": course_name,
                        "section": section,
                        "instructor": instructor,
                        "room": room,
                        "day": current_day,
                        "time_start": t_start,
                        "time_end": t_end,
                        "is_lab": is_lab,
                        "is_rescheduled": is_rescheduled,
                        "is_repeat": False,
                        "is_cancelled": is_cancelled,
                        "rag_summary": summary
                    })
    except Exception as e:
        logging.error(f"Error parsing FSE: {e}")
    return entries

def main():
    logging.info("Starting timetable fetch and parse process.")
    
    import os
    out_dir = "frontend/public"
    os.makedirs(out_dir, exist_ok=True)
    
    logging.info("Parsing FSC (School of Computing)...")
    fsc_entries = parse_fsc()
    with open(os.path.join(out_dir, "computing.json"), 'w', encoding='utf-8') as f:
        json.dump(fsc_entries, f, indent=2, ensure_ascii=False)
    logging.info(f"Saved {len(fsc_entries)} entries to computing.json")
    
    logging.info("Parsing FSM (School of Management)...")
    fsm_entries = parse_fsm()
    with open(os.path.join(out_dir, "management.json"), 'w', encoding='utf-8') as f:
        json.dump(fsm_entries, f, indent=2, ensure_ascii=False)
    logging.info(f"Saved {len(fsm_entries)} entries to management.json")
    
    logging.info("Parsing FSE (School of Engineering)...")
    fse_entries = parse_fse()
    with open(os.path.join(out_dir, "engineering.json"), 'w', encoding='utf-8') as f:
        json.dump(fse_entries, f, indent=2, ensure_ascii=False)
    logging.info(f"Saved {len(fse_entries)} entries to engineering.json")
        
    logging.info("Process completed successfully.")

if __name__ == "__main__":
    main()
